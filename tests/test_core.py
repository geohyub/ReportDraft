"""Tests for ProcessingReportDraft core — v3.0 expanded."""
import os
import sys
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from core import (
    parse_processing_log, generate_flow_from_template, generate_text_report,
    get_supported_types,
    ProcessingStep, ProcessingFlow,
    DEFAULT_SBP_STEPS, DEFAULT_UHR_STEPS,
    DEFAULT_MBES_STEPS, DEFAULT_MAG_STEPS, DEFAULT_SSS_STEPS,
    DATA_TYPE_TEMPLATES, SUPPORTED_DATA_TYPES,
    add_custom_step, remove_step, reorder_steps, update_step,
    compare_flows, FlowDiff,
    RevisionTracker, FlowRevision,
    PARAMETER_VALIDATION_RULES,
    validate_flow_parameters,
    get_flow_statistics,
    compare_flow_statistics,
    generate_excel_report,
    generate_json_export,
    generate_html_report,
    generate_all_templates,
    generate_bulk_docx,
)


# ── Template Tests ──
def test_default_sbp_steps():
    assert len(DEFAULT_SBP_STEPS) >= 8
    assert DEFAULT_SBP_STEPS[0].name == "Data Input"


def test_default_uhr_steps():
    assert len(DEFAULT_UHR_STEPS) >= 10
    assert DEFAULT_UHR_STEPS[0].name == "Data Input"


def test_default_mbes_steps():
    assert len(DEFAULT_MBES_STEPS) >= 8
    assert DEFAULT_MBES_STEPS[0].name == "Data Import"


def test_default_mag_steps():
    assert len(DEFAULT_MAG_STEPS) >= 8
    assert DEFAULT_MAG_STEPS[0].name == "Data Import"


def test_default_sss_steps():
    assert len(DEFAULT_SSS_STEPS) >= 8
    assert DEFAULT_SSS_STEPS[0].name == "Data Import"


def test_data_type_templates_dict():
    assert "SBP" in DATA_TYPE_TEMPLATES
    assert "UHR" in DATA_TYPE_TEMPLATES
    assert "MBES" in DATA_TYPE_TEMPLATES
    assert "MAG" in DATA_TYPE_TEMPLATES
    assert "SSS" in DATA_TYPE_TEMPLATES
    # MULTICHANNEL removed (duplicate of UHR)
    assert "2DHR" in DATA_TYPE_TEMPLATES


def test_supported_data_types():
    assert len(SUPPORTED_DATA_TYPES) >= 10
    assert "SBP" in SUPPORTED_DATA_TYPES
    assert "MBES" in SUPPORTED_DATA_TYPES


def test_template_sbp():
    flow = generate_flow_from_template("SBP")
    assert flow.data_type == "SBP"
    assert flow.software == "RadExPro"
    assert len(flow.steps) == len(DEFAULT_SBP_STEPS)


def test_template_uhr():
    flow = generate_flow_from_template("UHR")
    assert flow.data_type == "UHR"
    assert flow.software == "RadExPro"
    assert len(flow.steps) == len(DEFAULT_UHR_STEPS)


def test_template_mbes():
    flow = generate_flow_from_template("MBES")
    assert flow.data_type == "MBES"
    assert "CARIS" in flow.software
    assert len(flow.steps) == len(DEFAULT_MBES_STEPS)


def test_template_mag():
    flow = generate_flow_from_template("MAG")
    assert flow.data_type == "MAG"
    assert "Oasis" in flow.software
    assert len(flow.steps) == len(DEFAULT_MAG_STEPS)


def test_template_sss():
    flow = generate_flow_from_template("SSS")
    assert flow.data_type == "SSS"
    assert len(flow.steps) == len(DEFAULT_SSS_STEPS)


def test_template_2dhr():
    """2DHR is an alias for UHR (MULTICHANNEL was removed as duplicate)"""
    flow = generate_flow_from_template("2DHR")
    assert len(flow.steps) == len(DEFAULT_UHR_STEPS)


def test_template_sidescan():
    flow = generate_flow_from_template("SIDESCAN")
    assert len(flow.steps) == len(DEFAULT_SSS_STEPS)


def test_template_unknown_defaults_sbp():
    flow = generate_flow_from_template("UNKNOWN_TYPE")
    assert len(flow.steps) == len(DEFAULT_SBP_STEPS)


# ── Parser Tests ──
def test_parse_simple_log():
    log = """
Project: JAKO Offshore
Client: JAKO
Data type: SBP
Vessel: Geoview No.1
Software: RadExPro 2024.4

1. Data Input - Load SEG-Y files
   Format: SEG-Y Rev 1
   Byte order: Big Endian

2. Band-pass Filter - Remove noise
   Low cut: 200 Hz
   High cut: 4000 Hz

3. Gain - Apply AGC
   Window: 50 ms
"""
    flow = parse_processing_log(log)
    assert flow.project_name == "JAKO Offshore"
    assert flow.client == "JAKO"
    assert flow.data_type == "SBP"
    assert flow.vessel == "Geoview No.1"
    assert len(flow.steps) == 3
    assert flow.steps[0].name == "Data Input"
    assert flow.steps[1].parameters.get("Low cut") == "200 Hz"


def test_parse_numbered_steps():
    log = """
1) Import Data
2) Apply Filter
3) Stack
4) Export
"""
    flow = parse_processing_log(log)
    assert len(flow.steps) == 4
    assert flow.steps[2].name == "Stack"


def test_parse_with_parameters():
    log = """
1. Filter
   Type: Butterworth
   Order: 4
   Frequency: 100-2000 Hz
"""
    flow = parse_processing_log(log)
    assert len(flow.steps) == 1
    assert flow.steps[0].parameters["Type"] == "Butterworth"
    assert flow.steps[0].parameters["Order"] == "4"


def test_parse_empty():
    flow = parse_processing_log("")
    assert len(flow.steps) == 0
    assert flow.project_name == ""


def test_parse_metadata_only():
    log = """
Project: Test
Client: TestClient
Lines: 42
"""
    flow = parse_processing_log(log)
    assert flow.project_name == "Test"
    assert flow.client == "TestClient"
    assert flow.line_count == 42


def test_parse_area_and_vessel():
    log = "Area: East Sea\nVessel: Survey Ship Alpha"
    flow = parse_processing_log(log)
    assert flow.area == "East Sea"
    assert flow.vessel == "Survey Ship Alpha"


def test_parse_mbes_log():
    log = """
Project: BadaEnergy MBES Survey
Client: BadaEnergy
Data Type: MBES
Vessel: Geoview No.2

1. Import raw .all files
2. Apply SVP correction
3. Tide correction - Apply RTK tides
4. Swath editing
5. Surface generation
   Resolution: 1m
6. Export to BAG
"""
    flow = parse_processing_log(log)
    assert flow.project_name == "BadaEnergy MBES Survey"
    assert flow.data_type == "MBES"
    assert len(flow.steps) == 6


# ── Dataclass Tests ──
def test_processing_step_dataclass():
    step = ProcessingStep(order=1, name="Filter", description="Apply bandpass", parameters={"freq": "100 Hz"})
    assert step.order == 1
    assert step.parameters["freq"] == "100 Hz"


def test_processing_flow_step_count():
    flow = ProcessingFlow()
    flow.steps = [ProcessingStep(i, f"Step {i}") for i in range(5)]
    assert flow.step_count == 5


def test_processing_flow_defaults():
    flow = ProcessingFlow()
    assert flow.data_type == "SBP"
    assert flow.software == "RadExPro"
    assert flow.step_count == 0


# ── Text Report Tests ──
def test_text_report_basic():
    flow = generate_flow_from_template("SBP")
    flow.project_name = "Test Project"
    report = generate_text_report(flow)
    assert "ProcessingReportDraft" in report
    assert "Test Project" in report
    assert "Step 1" in report


def test_text_report_mbes():
    flow = generate_flow_from_template("MBES")
    flow.project_name = "MBES Survey"
    report = generate_text_report(flow)
    assert "MBES Survey" in report
    assert "Data Import" in report


def test_text_report_empty_flow():
    flow = ProcessingFlow()
    report = generate_text_report(flow)
    assert "ProcessingReportDraft" in report
    assert "0 steps" in report


def test_text_report_with_notes():
    flow = ProcessingFlow(notes="Check frequency content carefully")
    report = generate_text_report(flow)
    assert "Check frequency content" in report


# ── Supported Types Tests ──
def test_get_supported_types():
    info = get_supported_types()
    assert len(info) >= 5
    assert "SBP" in info
    assert "step_count" in info["SBP"]
    assert info["SBP"]["step_count"] == len(DEFAULT_SBP_STEPS)


def test_supported_types_preview():
    info = get_supported_types()
    assert "steps_preview" in info["MBES"]
    assert len(info["MBES"]["steps_preview"]) <= 5


# ── DOCX Generation Tests ──
def test_generate_docx(tmp_path):
    from core import generate_docx_report
    flow = generate_flow_from_template("SBP")
    flow.project_name = "Test Project"
    flow.client = "Test Client"
    out = str(tmp_path / "report.docx")
    generate_docx_report(flow, out)
    assert os.path.exists(out)
    assert os.path.getsize(out) > 1000


def test_generate_docx_from_parsed(tmp_path):
    from core import generate_docx_report
    log = """
Project: JAKO
Client: JAKO Corp
1. Input - Load data
2. Filter - Bandpass
   Low: 100 Hz
   High: 3000 Hz
3. Output
"""
    flow = parse_processing_log(log)
    out = str(tmp_path / "parsed_report.docx")
    generate_docx_report(flow, out)
    assert os.path.exists(out)


def test_generate_docx_empty_flow(tmp_path):
    from core import generate_docx_report
    flow = ProcessingFlow()
    out = str(tmp_path / "empty.docx")
    generate_docx_report(flow, out)
    assert os.path.exists(out)


def test_generate_docx_mbes(tmp_path):
    from core import generate_docx_report
    flow = generate_flow_from_template("MBES")
    flow.project_name = "MBES Project"
    out = str(tmp_path / "mbes_report.docx")
    generate_docx_report(flow, out)
    assert os.path.exists(out)
    assert os.path.getsize(out) > 1000


def test_generate_docx_mag(tmp_path):
    from core import generate_docx_report
    flow = generate_flow_from_template("MAG")
    flow.project_name = "MAG Survey"
    out = str(tmp_path / "mag_report.docx")
    generate_docx_report(flow, out)
    assert os.path.exists(out)


# ── Step Parameter Tests ──
def test_sbp_steps_have_parameters():
    for step in DEFAULT_SBP_STEPS:
        if step.name != "Trace Editing":
            assert len(step.parameters) > 0, f"Step '{step.name}' should have parameters"


def test_mbes_steps_have_descriptions():
    for step in DEFAULT_MBES_STEPS:
        assert step.description, f"MBES step '{step.name}' should have description"


def test_mag_igrf_step():
    igrf_steps = [s for s in DEFAULT_MAG_STEPS if "IGRF" in s.name]
    assert len(igrf_steps) == 1
    assert "IGRF-13" in igrf_steps[0].parameters.get("IGRF model", "")


def test_sss_mosaicking_step():
    mosaic_steps = [s for s in DEFAULT_SSS_STEPS if "Mosaic" in s.name]
    assert len(mosaic_steps) == 1
    assert "Resolution" in mosaic_steps[0].parameters


# ── Custom Step Editing Tests ──

def test_add_custom_step_append():
    flow = generate_flow_from_template("SBP")
    original_count = len(flow.steps)
    add_custom_step(flow, "Custom QC", "Run custom QC", {"method": "visual"})
    assert len(flow.steps) == original_count + 1
    assert flow.steps[-1].name == "Custom QC"
    assert flow.steps[-1].order == original_count + 1
    assert flow.steps[-1].parameters == {"method": "visual"}


def test_add_custom_step_at_position():
    flow = generate_flow_from_template("SBP")
    original_first = flow.steps[0].name
    add_custom_step(flow, "Pre-check", "Initial validation", position=1)
    assert flow.steps[0].name == "Pre-check"
    assert flow.steps[0].order == 1
    assert flow.steps[1].name == original_first
    assert flow.steps[1].order == 2


def test_remove_step():
    flow = generate_flow_from_template("SBP")
    original_count = len(flow.steps)
    remove_step(flow, 1)
    assert len(flow.steps) == original_count - 1


def test_remove_step_renumbers():
    flow = ProcessingFlow()
    flow.steps = [
        ProcessingStep(1, "A"),
        ProcessingStep(2, "B"),
        ProcessingStep(3, "C"),
    ]
    remove_step(flow, 2)
    assert len(flow.steps) == 2
    assert flow.steps[0].order == 1
    assert flow.steps[0].name == "A"
    assert flow.steps[1].order == 2
    assert flow.steps[1].name == "C"


def test_reorder_steps():
    flow = ProcessingFlow()
    flow.steps = [
        ProcessingStep(1, "A"),
        ProcessingStep(2, "B"),
        ProcessingStep(3, "C"),
    ]
    reorder_steps(flow, [3, 1, 2])
    assert flow.steps[0].name == "C"
    assert flow.steps[1].name == "A"
    assert flow.steps[2].name == "B"
    # Orders should be renumbered
    assert flow.steps[0].order == 1
    assert flow.steps[1].order == 2
    assert flow.steps[2].order == 3


def test_reorder_steps_invalid():
    flow = ProcessingFlow()
    flow.steps = [
        ProcessingStep(1, "A"),
        ProcessingStep(2, "B"),
    ]
    with pytest.raises(ValueError):
        reorder_steps(flow, [1, 2, 99])


def test_update_step_name():
    flow = ProcessingFlow()
    flow.steps = [ProcessingStep(1, "Old Name", "desc")]
    update_step(flow, 1, name="New Name")
    assert flow.steps[0].name == "New Name"
    assert flow.steps[0].description == "desc"  # unchanged


def test_update_step_parameters():
    flow = ProcessingFlow()
    flow.steps = [ProcessingStep(1, "Filter", "desc", {"freq": "100"})]
    update_step(flow, 1, parameters={"freq": "200", "type": "Butterworth"})
    assert flow.steps[0].parameters == {"freq": "200", "type": "Butterworth"}
    assert flow.steps[0].name == "Filter"  # unchanged


# ── Flow Comparison Tests ──

def test_compare_flows_identical():
    flow1 = generate_flow_from_template("SBP")
    flow2 = generate_flow_from_template("SBP")
    diff = compare_flows(flow1, flow2)
    assert len(diff.added_steps) == 0
    assert len(diff.removed_steps) == 0
    assert len(diff.modified_steps) == 0
    assert len(diff.metadata_changes) == 0


def test_compare_flows_added_steps():
    flow1 = generate_flow_from_template("SBP")
    flow2 = generate_flow_from_template("SBP")
    add_custom_step(flow2, "New Custom Step", "description")
    diff = compare_flows(flow1, flow2)
    assert len(diff.added_steps) == 1
    assert diff.added_steps[0].name == "New Custom Step"


def test_compare_flows_removed_steps():
    flow1 = generate_flow_from_template("SBP")
    flow2 = generate_flow_from_template("SBP")
    removed_name = flow2.steps[0].name
    remove_step(flow2, 1)
    diff = compare_flows(flow1, flow2)
    assert len(diff.removed_steps) == 1
    assert diff.removed_steps[0].name == removed_name


def test_compare_flows_modified_parameters():
    flow1 = generate_flow_from_template("SBP")
    flow2 = generate_flow_from_template("SBP")
    update_step(flow2, 1, parameters={"Input format": "SEG-D"})
    diff = compare_flows(flow1, flow2)
    assert len(diff.modified_steps) == 1
    assert diff.modified_steps[0]["name"] == "Data Input"


def test_compare_flows_metadata_changes():
    flow1 = ProcessingFlow(project_name="Project A", client="Client A")
    flow2 = ProcessingFlow(project_name="Project B", client="Client B")
    diff = compare_flows(flow1, flow2)
    assert "project_name" in diff.metadata_changes
    assert diff.metadata_changes["project_name"]["old"] == "Project A"
    assert diff.metadata_changes["project_name"]["new"] == "Project B"
    assert "client" in diff.metadata_changes


# ── Revision Tracking Tests ──

def test_revision_tracker_save():
    tracker = RevisionTracker()
    flow = generate_flow_from_template("SBP")
    flow.project_name = "Test Project"
    rev = tracker.save_revision(flow, author="Kim", changes="Initial draft")
    assert rev.version == 1
    assert rev.author == "Kim"
    assert rev.changes == "Initial draft"
    assert rev.flow_snapshot["project_name"] == "Test Project"


def test_revision_tracker_history():
    tracker = RevisionTracker()
    flow = generate_flow_from_template("SBP")
    tracker.save_revision(flow, author="Kim", changes="v1")
    add_custom_step(flow, "Extra Step")
    tracker.save_revision(flow, author="Park", changes="v2")
    history = tracker.get_history()
    assert len(history) == 2
    assert history[0]["version"] == 1
    assert history[0]["author"] == "Kim"
    assert history[1]["version"] == 2
    assert history[1]["author"] == "Park"


def test_revision_tracker_diff():
    tracker = RevisionTracker()
    flow = generate_flow_from_template("SBP")
    flow.project_name = "Project v1"
    tracker.save_revision(flow, author="Kim", changes="Initial")
    flow.project_name = "Project v2"
    add_custom_step(flow, "Extra Step")
    tracker.save_revision(flow, author="Kim", changes="Added extra step")
    diff = tracker.diff_revisions(1, 2)
    assert "project_name" in diff.metadata_changes
    assert len(diff.added_steps) == 1
    assert diff.added_steps[0].name == "Extra Step"


# ── Parameter Validation Tests ──

def test_validate_params_sbp_valid():
    flow = generate_flow_from_template("SBP")
    result = validate_flow_parameters(flow)
    assert result["total_params_checked"] > 0
    assert result["score"] >= 0


def test_validate_params_empty_flow():
    flow = ProcessingFlow(data_type="SBP", steps=[])
    result = validate_flow_parameters(flow)
    assert result["total_params_checked"] == 0
    assert result["score"] == 100.0


def test_validate_params_invalid_value():
    flow = generate_flow_from_template("SBP")
    # Add a step with invalid parameter
    add_custom_step(flow, "Band-pass Filter", parameters={"Low Cut (Hz)": -999})
    result = validate_flow_parameters(flow)
    assert result["invalid"] >= 1
    assert len(result["issues"]) >= 1


def test_validate_params_unknown_type():
    flow = ProcessingFlow(data_type="UNKNOWN", steps=[
        ProcessingStep(order=1, name="Test", parameters={"x": "y"})
    ])
    result = validate_flow_parameters(flow)
    assert result["unknown"] >= 1


def test_validate_params_choice_valid():
    flow = ProcessingFlow(data_type="SBP", steps=[
        ProcessingStep(order=1, name="Band-pass Filter", parameters={"Filter Type": "Butterworth"})
    ])
    result = validate_flow_parameters(flow)
    assert result["valid"] >= 1
    assert result["invalid"] == 0


def test_validate_params_choice_invalid():
    flow = ProcessingFlow(data_type="SBP", steps=[
        ProcessingStep(order=1, name="Band-pass Filter", parameters={"Filter Type": "InvalidType"})
    ])
    result = validate_flow_parameters(flow)
    assert result["invalid"] >= 1
    assert any(i["severity"] == "error" for i in result["issues"])


def test_validate_params_non_numeric():
    flow = ProcessingFlow(data_type="SBP", steps=[
        ProcessingStep(order=1, name="Band-pass Filter", parameters={"Low Cut (Hz)": "not_a_number"})
    ])
    result = validate_flow_parameters(flow)
    assert result["invalid"] >= 1
    assert any(i["severity"] == "error" for i in result["issues"])


# ── Flow Statistics Tests ──

def test_flow_statistics_sbp():
    flow = generate_flow_from_template("SBP")
    stats = get_flow_statistics(flow)
    assert stats["step_count"] >= 8
    assert stats["total_parameters"] > 0
    assert stats["data_type"] == "SBP"


def test_flow_statistics_empty():
    flow = ProcessingFlow(data_type="SBP", steps=[])
    stats = get_flow_statistics(flow)
    assert stats["step_count"] == 0
    assert stats["completeness_score"] == 0.0


def test_flow_statistics_has_metadata():
    flow = generate_flow_from_template("SBP")
    flow.project_name = "Test"
    flow.client = "Client"
    flow.vessel = "Vessel"
    stats = get_flow_statistics(flow)
    assert stats["has_metadata"] is True


def test_flow_statistics_no_metadata():
    flow = ProcessingFlow(data_type="SBP", steps=[])
    stats = get_flow_statistics(flow)
    assert stats["has_metadata"] is False


def test_compare_flow_statistics():
    flows = [generate_flow_from_template(t) for t in ["SBP", "UHR", "MBES"]]
    result = compare_flow_statistics(flows)
    assert result["flow_count"] == 3
    assert len(result["comparison"]) == 3
    assert result["avg_step_count"] > 0


def test_compare_flow_statistics_single():
    flows = [generate_flow_from_template("SBP")]
    result = compare_flow_statistics(flows)
    assert result["flow_count"] == 1
    assert result["most_detailed"] == "SBP"
    assert result["least_detailed"] == "SBP"


# ── Excel Export Tests ──

def test_generate_excel_report(tmp_path):
    flow = generate_flow_from_template("SBP")
    flow.project_name = "TestProject"
    output = str(tmp_path / "report.xlsx")
    result = generate_excel_report(flow, output)
    assert os.path.exists(result)
    import openpyxl
    wb = openpyxl.load_workbook(output)
    assert "Overview" in wb.sheetnames
    assert "Processing Steps" in wb.sheetnames
    assert "Validation" in wb.sheetnames


def test_generate_excel_report_empty_flow(tmp_path):
    flow = ProcessingFlow(data_type="SBP")
    output = str(tmp_path / "empty_report.xlsx")
    result = generate_excel_report(flow, output)
    assert os.path.exists(result)


# ── JSON Export Tests ──

def test_generate_json_export():
    flow = generate_flow_from_template("SBP")
    json_str = generate_json_export(flow)
    import json
    data = json.loads(json_str)
    assert data["data_type"] == "SBP"
    assert len(data["steps"]) >= 8


def test_generate_json_export_roundtrip():
    flow = generate_flow_from_template("MAG")
    flow.project_name = "JSON Test"
    import json
    json_str = generate_json_export(flow)
    data = json.loads(json_str)
    assert data["project_name"] == "JSON Test"
    assert data["data_type"] == "MAG"
    assert data["step_count"] == len(data["steps"])


# ── HTML Export Tests ──

def test_generate_html_report():
    flow = generate_flow_from_template("SBP")
    html = generate_html_report(flow)
    assert "<html" in html
    assert "SBP" in html
    assert "Processing" in html


def test_generate_html_report_has_steps():
    flow = generate_flow_from_template("MBES")
    html = generate_html_report(flow)
    assert "Data Import" in html
    assert "step-card" in html


# ── Bulk Template Tests ──

def test_generate_all_templates():
    result = generate_all_templates(project_name="TestProj")
    assert len(result) == 5
    assert "SBP" in result
    assert "UHR" in result
    assert "MBES" in result
    assert "MAG" in result
    assert "SSS" in result
    assert result["SBP"].project_name == "TestProj"


def test_generate_all_templates_metadata():
    result = generate_all_templates(
        project_name="Proj", client="Client", vessel="V1", area="Area1"
    )
    for dt, flow in result.items():
        assert flow.project_name == "Proj"
        assert flow.client == "Client"
        assert flow.vessel == "V1"
        assert flow.area == "Area1"


def test_generate_bulk_docx(tmp_path):
    flows = generate_all_templates()
    output_dir = str(tmp_path)
    results = generate_bulk_docx(flows, output_dir)
    assert len(results) == 5
    for path in results:
        assert os.path.exists(path)
